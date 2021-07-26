from docxtpl import DocxTemplate
from num2words import num2words
from openpyxl import load_workbook
import os
from django.core.files import File

from Daylesford.settings import BASE_DIR
from main import models
from main.utils import save_doc_after_generation

def gen_doc(offer, company, product, deal):
    # Документ 2
    # name_of_contract = '113/ДМА'
    # name_of_provider = 'Общество с ограниченной ответственностью "Маныч-Агро"'
    # head_of_provider = 'Директора Ротко Анатолия Викторовича'
    # basis_of_doc = 'Устава'
    # name_of_product = 'КУКУРУЗА'
    # harvest_year = '2020'
    # address_load = '344002, г. Ростов-на-Дону, ул. 1-я Луговая, 42'
    # volume_product = '2 200'
    # volume_product_in_words = 'Две тысячи двести'
    volume_product_in_words = num2words(2200, lang='ru') + ' рублей'
    sum_per_tonne = '18 150,00'
    sum_per_tonne_in_words = 'Восемнадцать тысяч сто пятьдесят рублей 00 копеек'
    sum_per_tonne_in_words = num2words(18150, lang='ru') + ' рублей 00 копеек'
    # date_start_shipment = '29 января 2021 г.'
    # date_finish_shipment = '28 февраля 2021 г.'
    size_NDS_per_tonne = '1 650,00'
    # total_sum = '39 930 000,00'
    # total_sum_in_words = 'Тридцать девять миллионов девятьсот тридцать тысяч рублей 00 копеек'
    total_sum_in_words = num2words(39930000, lang='ru') + ' рублей 00 копеек'
    size_NDS = '3 630 000,00'
    # number_of_spec = '1'
    # date_of_contract = '27 января 2021 г.'
    # date_of_spec = '28 января 2021 г.'
    TEXT1 = 'Датой поставки считается дата передачи Товара уполномоченному представителю Покупателя'
    TEXT2 = ' Приёмка Товара по качеству и количеству осуществляется при погрузке товара в автотранспорт Покупателя. ' \
            'После приемки Товара по количеству подписывается товарно-транспортная накладная СП-31, по утвержденной ' \
            'в Приложении к настоящей Спецификации форме'

    # Документ 1
    # date_finish_of_contract = '31 декабря 2021 г.'
    # ul_address = '346601, Ростовская обл, Багаевский р-н, Манычская ст-ца, Магаданская ул, дом 7'
    # inn = '6166018099'
    # kpp = '610301001'
    # phone_number = '8-86357-43-2-33'
    # email_of_head = 'manychagro@mail.ru'
    # ogrn = '1026103161336'
    # payment_account = '40702810900900000450'
    # name_of_bank = 'ЮГ-ИНВЕСТБАНК (ПАО) г. Краснодар'
    # correspondent_account = '30101810600000000966'
    # bik = '040349966'
    # position_head_of_provider = 'Директор'
    # short_fio = 'Ротко А.В.'

    doc = DocxTemplate("./temp_file/specification.docx")
    context = {
        'number_of_spec': deal.number_of_spec,
        'name_of_contract': deal.name_of_contract,
        'date_of_contract': deal.date_start_of_contract,
        'date_of_spec': deal.date_start_of_spec,
        'name_of_provider': company.name_of_provider,
        'head_of_provider': company.head_of_provider,
        'basis_of_doc': company.basis_of_doc,
        'name_of_product': product.title,
        'harvest_year': product.harvest_year,
        'volume_product': offer.volume,
        'volume_product_in_words': volume_product_in_words,
        'date_start_shipment': deal.date_start_shipment,
        'date_finish_shipment': deal.date_finish_shipment,
        'sum_per_tonne': sum_per_tonne,
        'sum_per_tonne_in_words': sum_per_tonne_in_words,
        'size_NDS_per_tonne': size_NDS_per_tonne,
        'total_sum': offer.cost,
        'total_sum_in_words': total_sum_in_words,
        'size_NDS': size_NDS,
        'address_load': company.address_load,
        'TEXT1': TEXT1,
        'TEXT2': TEXT2,
        'date_finish_of_contract': deal.date_finish_of_contract,
        'ul_address': company.ul_address,
        'inn': company.inn,
        'kpp': company.kpp,
        'phone_number': company.phone_number,
        'email_of_head': company.email_of_head,
        'ogrn': company.ogrn,
        'payment_account': company.payment_account,
        'name_of_bank': company.name_of_bank,
        'correspondent_account': company.correspondent_account,
        'bik': company.bik,
        'position_head_of_provider': company.position_head_of_provider,
        'short_fio': company.short_fio,
    }
    # doc.render(context)
    # doc.save("generated.docx")
    # import pythoncom
    # pythoncom.CoInitializeEx(0)
    # from docx2pdf import convert
    #
    # # convert("generated.docx")
    # convert("generated.docx", "output.pdf")
    # # convert("my_docx_folder/")

#
# {
#   "name_of_provider": "Общество с ограниченной ответственностью 'Маныч-Агро'",
#   "head_of_provider": "Ротко Анатолия Викторовича",
#   "short_fio": "Ротко А.В.",
#   "position_head_of_provider": "Директор",
#   "basis_of_doc": "Устава",
#   "address_load": "344002. г. Ростов-на-Дону. ул. 1-я Луговая 42",
#   "ul_address": "346601. Ростовская обл. Багаевский р-н. Манычская ст-ца. Магаданская ул дом 7",
#   "inn": 6166018099,
#   "kpp": 610301001,
#   "ogrn": 1026103161336,
#   "bik": 040349966,
#   "payment_account": "40702810900900000450",
#   "correspondent_account": "30101810600000000966",
#   "phone_number": "8-86357-43-2-33",
#   "email_of_head": "manychagro@mail.ru",
#   "name_of_bank": "ЮГ-ИНВЕСТБАНК (ПАО) г. Краснодар"
# }

def gen_doc_specification():
    tpl_path = os.path.join(BASE_DIR, 'temp_file' , 'template_doc', 'specification.docx')
    tpl = DocxTemplate(tpl_path)


    name_of_contract = '113/ДМА'
    name_of_provider = 'Общество с ограниченной ответственностью "Маныч-Агро"'
    head_of_provider = 'Директора Ротко Анатолия Викторовича'
    basis_of_doc = 'Устава'
    name_of_product = 'КУКУРУЗА'
    harvest_year = '2020'
    address_load = '344002, г. Ростов-на-Дону, ул. 1-я Луговая, 42'
    volume_product = '2 200'
    volume_product_in_words = 'Две тысячи двести'
    volume_product_in_words = num2words(2200, lang='ru') + ' рублей'
    sum_per_tonne = '18 150,00'
    sum_per_tonne_in_words = 'Восемнадцать тысяч сто пятьдесят рублей 00 копеек'
    sum_per_tonne_in_words = num2words(18150, lang='ru') + ' рублей 00 копеек'
    date_start_of_contract = '29 января 2021 г.'
    date_start_of_spec = '29 января 2021 г.'
    date_start_shipment = '29 января 2021 г.'
    date_finish_shipment = '28 февраля 2021 г.'
    size_NDS_per_tonne = '1 650,00'
    total_sum = '39 930 000,00'
    total_sum_in_words = 'Тридцать девять миллионов девятьсот тридцать тысяч рублей 00 копеек'
    total_sum_in_words = num2words(39930000, lang='ru') + ' рублей 00 копеек'
    size_NDS = '3 630 000,00'
    number_of_spec = '1'
    date_of_contract = '27 января 2021 г.'
    date_of_spec = '28 января 2021 г.'
    TEXT1 = 'Датой поставки считается дата передачи Товара уполномоченному представителю Покупателя'
    TEXT2 = ' Приёмка Товара по качеству и количеству осуществляется при погрузке товара в автотранспорт Покупателя. ' \
            'После приемки Товара по количеству подписывается товарно-транспортная накладная СП-31, по утвержденной ' \
            'в Приложении к настоящей Спецификации форме'

    context = {
        'number_of_spec': number_of_spec,
        'name_of_contract': name_of_contract,
        'date_of_contract': date_start_of_contract,
        'date_of_spec': date_start_of_spec,
        'name_of_provider': name_of_provider,
        'head_of_provider': head_of_provider,
        'basis_of_doc': basis_of_doc,
        'name_of_product': name_of_product,
        'harvest_year': harvest_year,
        'volume_product': volume_product,
        'volume_product_in_words': volume_product_in_words,
        'date_start_shipment': date_start_shipment,
        'date_finish_shipment': date_finish_shipment,
        'sum_per_tonne': sum_per_tonne,
        'sum_per_tonne_in_words': sum_per_tonne_in_words,
        'size_NDS_per_tonne': size_NDS_per_tonne,
        'total_sum': total_sum,
        'total_sum_in_words': total_sum_in_words,
        'size_NDS': size_NDS,
        'address_load': address_load,
        'TEXT1': TEXT1,
        'TEXT2': TEXT2,
    }

    tpl.render(context)
    tpl_save_path = os.path.join(BASE_DIR, 'temp_file', 'temporary_files', 'Спецификация.docx')
    tpl.save(tpl_save_path)
    save_doc_after_generation('Спецификация.docx', 'specification', tpl_save_path)


def gen_doc_contract():
    tpl_path = os.path.join(BASE_DIR, 'temp_file' , 'template_doc', 'contract.docx')
    tpl = DocxTemplate(tpl_path)

    name_of_contract = '113/ДМА'
    date_start_of_contract = '27 января 2021 г.'
    name_of_provider = 'Общество с ограниченной ответственностью "Маныч-Агро"'
    head_of_provider = 'Директора Ротко Анатолия Викторовича'
    basis_of_doc = "Устава"
    date_finish_of_contract = '31 декабря 2021 г.'
    ul_address = '346601, Ростовская обл, Багаевский р-н, Манычская ст-ца, Магаданская ул, дом 7'
    inn = '6166018099'
    kpp = '610301001'
    phone_number = '8-86357-43-2-33'
    email_of_head = 'manychagro@mail.ru'
    ogrn = '1026103161336'
    payment_account = '40702810900900000450'
    name_of_bank = 'ЮГ-ИНВЕСТБАНК (ПАО) г. Краснодар'
    correspondent_account = '30101810600000000966'
    bik = '040349966'
    position_head_of_provider = 'Директор'
    short_fio = 'Ротко А.В.'

    context = {
        'name_of_contract': name_of_contract,
        'date_of_contract': date_start_of_contract,
        'name_of_provider': name_of_provider,
        'head_of_provider': head_of_provider,
        'basis_of_doc': basis_of_doc,
        'date_finish_of_contract': date_finish_of_contract,
        'ul_address': ul_address,
        'inn': inn,
        'kpp': kpp,
        'phone_number': phone_number,
        'email_of_head': email_of_head,
        'ogrn': ogrn,
        'payment_account': payment_account,
        'name_of_bank': name_of_bank,
        'correspondent_account': correspondent_account,
        'bik': bik,
        'position_head_of_provider': position_head_of_provider,
        'short_fio': short_fio,
    }

    tpl.render(context)
    tpl_save_path = os.path.join(BASE_DIR, 'temp_file', 'temporary_files', 'Договор.docx')
    tpl.save(tpl_save_path)
    save_doc_after_generation('Договор.docx', 'contract', tpl_save_path)

def gen_doc_payment_invoice():
    wb = load_workbook(filename='backend/temp_file/init_doc/Счет на оплату_sample (1).xlsx')
    ws = wb.active

    recipients_bank = 'ЦЕНТРАЛЬНО-ЧЕРНОЗЕМНЫЙ БАНК ПАО СБЕРБАНК г. Воронеж' # B3
    bik = '042007681' # AO3
    payment_account_1 = '30101810600000000681' # AO4
    payment_account = '40702810513230102419' # AO6
    inn = '3603005653' # E6
    kpp = '360301001' # S6
    recipient = 'Сельскохозяйственная артель (колхоз) "РОДИНА"' # B7
    payment_invoice_number = 'Счет на оплату № 1 от 10 января 2020 г.' # B11
    provider = 'Сельскохозяйственная артель (колхоз) "РОДИНА", ИНН 3603005653, ' \
      'КПП 360301001, 396755, Воронежская обл, Богучарский р-н, Данцевка с, ' \
      'Новая ул, дом 11, тел.: +7 989 538-66-38 (агент)' # G15
    reason_doc = '№ 99 от 25.12.2019 (руб.)' # G21
    goods = 'ПОДСОЛНЕЧНИК' # D24
    quantity = '18,55' # AJ24
    cost = '17 700,00' # AS24
    total_cost = '328 335,00' # AX24, AY26, AY28
    result = 'Всего наименований 1, на сумму 328 335,00 RUB' # B29
    total_cost_in_words = 'Триста двадцать восемь тысяч триста тридцать пять рублей 00 копеек' # B30

    ws['B3'] = recipients_bank
    ws['AO3'] = bik
    ws['AO4'] = payment_account_1
    ws['AO6'] = payment_account
    ws['E6'] = inn
    ws['S6'] = kpp
    ws['B11'] = payment_invoice_number
    ws['B7'] = recipient
    ws['G15'] = provider
    ws['G21'] = reason_doc
    ws['D24'] = goods
    ws['AJ24'] = quantity
    ws['AS24'] = cost
    ws['AX24'] = total_cost
    ws['AY26'] = total_cost
    ws['AY28'] = total_cost
    ws['B29'] = result
    ws['B30'] = total_cost_in_words

    wb.save('12.xlsx')

def gen_doc_loading_register():
    tpl_path = os.path.join(BASE_DIR, 'temp_file' , 'template_doc', 'loading_register.docx')
    tpl = DocxTemplate(tpl_path)

    context = {
        'company': 'Общество с ограниченной ответственностью «Дейлсфорд Мёрчант»',
        'ogrn': '1167746749664',
        'inn_and_kpp': '7707371041/772701001',
        'address': '117447, г. Москва, ул. Черемушкинская  Б., дом  13, строение 4, пом II ком 25',
        'product': 'кукурузы',
        'provider': 'КАЗЬМИНСКИЙ СПК КОЛХОЗ-ПЛЕМЗАВОД',
        'date': '17.07.2021',
        'delivery_address': 'Ставропольский край, Кочубеевский р-он, Казьминское с, Советская ул, д 48',
        'tbl_contents': [
            {'car': 'В179ке126', 'trailer': 'ха9379-26', 'driver': 'Ивенский С.Н',
             'company': 'ИП Таран С.В., ИНН 231513771277,  Краснодарский кр., Новороссийск г., '
                        'ул. Энгельса , д. 84, кв. 56'},
        ]
    }

    tpl.render(context)
    tpl_save_path = os.path.join(BASE_DIR, 'temp_file', 'temporary_files', 'Реестр_погрузки.docx')
    tpl.save(tpl_save_path)
    save_doc_after_generation('Реестр_погрузки.docx', 'loading_plan', tpl_save_path)

def gen_doc_act_check():
    tpl_path = os.path.join(BASE_DIR, 'temp_file', 'init_doc', 'Акт сверки взаиморасчетов_sample (1).xlsx')
    wb = load_workbook(filename=tpl_path)
    ws = wb.active

    date_start = '01.01.2021'
    date_finish = '20.07.2021'

    customer = 'ООО "ДЕЙЛСФОРД МЁРЧАНТ"'
    customer_inn = '7707371041'
    provider = 'ИП Лесных Виктор Николаевич'
    provider_inn = '360600037753'
    contract = 'договору № 53/ДМ от 04.08.2020 (RUB)'

    text_for_B3_to_P3 = 'взаимных расчетов за период: {0} - {1} \n между {2} (ИНН {3}) \n' \
                        'и {4} (ИНН {5}) \n по {6}'.format(
        date_start, date_finish, customer, customer_inn, provider, provider_inn, contract
    ) # B3:P3

    customer_fio = 'Шевцова Екатерина Юрьевна'
    customer_position = 'Старший бухгалтер'
    customer_fio_short = 'Шевцова Е. Ю.' # D22:G22

    provider_fio = 'ИП Лесных Виктор Николаевич'

    text_for_B5_to_P5 = 'Мы, нижеподписавшиеся, {0} {1} ' \
                        '{2}, с одной стороны, и {3}, ' \
                        'с другой стороны, составили настоящий акт сверки в том, что состояние взаимных ' \
                        'расчетов по данным учета следующее:'.format(
        customer_position, customer, customer_fio, provider_fio
    ) # B5:P5

    text_for_B7_to_I7 = 'По данным {0},  RUB'.format(customer) # B7:I7
    text_for_J7_to_P7 = 'По данным {0},  RUB'.format(provider) # J7:P7

    customer_balance_date1 = '14.07.21' # B10
    customer_balance_date2 = '15.07.21' # B11
    customer_balance_doc1 = 'Оплата (2172 от 14.07.2021)' # C10:D10
    customer_balance_doc2 = 'Приход (1 от 15.07.2021)' # C11:D11
    customer_debit = '3472000' # E10:F10
    customer_credit = '2869112' # G11:I11
    customer_debit_total = '3472000' # E12:F12
    customer_credit_total = '2869112' # G12:I12
    customer_debit_balance_total = '602888' # E13:F13

    provider_balance_date1 = '14.07.21' # J10
    provider_balance_date2 = '15.07.21' # J11
    provider_balance_doc1 = 'Оплата (2172 от 14.07.2021)' # K10:L10
    provider_balance_doc2 = 'Приход (1 от 15.07.2021)' # K11:L11
    provider_debit = '2869112' # M11:N11
    provider_credit = '3472000' # O10:P10
    provider_debit_total = '2869112' # M12:N12
    provider_credit_total = '3472000' # O12:P12
    provider_credit_balance_total = '602888' # O13:P13

    text_for_B15_to_G15 = 'По данным {0}'.format(customer) # B15:G15
    text_for_I15_to_O15 = 'По данным {0}'.format(provider) # I15:O15

    total_date = '20.07.2021'
    total_cost = '602 888,00'
    total_cost_in_words = 'Шестьсот две тысячи восемьсот восемьдесят восемь рублей 00 копеек'
    total_text_customer = 'на {0} задолженность в пользу {1}' \
                          ' {2} RUB ({3}).'.format(total_date, customer, total_cost, total_cost_in_words) # B16:G16, I16:O16

    text_for_B18_to_G18 = 'От {0}'.format(customer) # B18:G18
    text_for_I18_to_O18 = 'От {0}'.format(provider) # I18:O18


    ws['B3'] = text_for_B3_to_P3
    ws['B5'] = text_for_B5_to_P5
    ws['B7'] = text_for_B7_to_I7
    ws['J7'] = text_for_J7_to_P7
    ws['B10'] = customer_balance_date1
    ws['B11'] = customer_balance_date2
    ws['C10'] = provider_balance_doc1
    ws['C11'] = provider_balance_doc2
    ws['E10'] = customer_debit
    ws['G11'] = customer_credit
    ws['E12'] = customer_debit_total
    ws['G12'] = customer_credit_total
    ws['E13'] = customer_debit_balance_total
    ws['J10'] = provider_balance_date1
    ws['J11'] = provider_balance_date2
    ws['K10'] = provider_balance_doc1
    ws['K11'] = provider_balance_doc2
    ws['M11'] = provider_debit
    ws['O10'] = provider_credit
    ws['M12'] = provider_debit_total
    ws['O12'] = provider_credit_total
    ws['O13'] = provider_credit_balance_total
    ws['B15'] = text_for_B15_to_G15
    ws['I15'] = text_for_I15_to_O15
    ws['B16'] = total_text_customer
    ws['I16'] = total_text_customer
    ws['B18'] = text_for_B18_to_G18
    ws['I18'] = text_for_I18_to_O18
    ws['B20'] = customer_position
    ws['D22'] = customer_fio_short

    tpl_save_path = os.path.join(BASE_DIR, 'temp_file', 'temporary_files', 'Акт_сверки_взаиморасчетов.xlsx')
    wb.save(tpl_save_path)
    save_doc_after_generation('Акт_сверки_взаиморасчетов.xlsx', 'verification_act', tpl_save_path)

def gen_doc_YPD():
    tpl_path = os.path.join(BASE_DIR, 'temp_file', 'init_doc', 'УПД (статус 1)_sample.xlsx')
    wb = load_workbook(filename=tpl_path)
    ws = wb.active

    invoice_number = '579' # P1:U1
    invoice_date = '12 июля 2021 г.' # Y1:AF1

    provider = 'Общество с ограниченной ответственностью "СИТИТОРГ-АГРО"' # R4:AV4
    provider_address = '397702, Воронежская обл, Бобровский р-н, Бобров г, Им.Котовского ул, дом 2' # R5:AV5
    provider_inn_and_kpp = '3663103831/360201001' # R6:AV6
    text_for_R7_to_AV7 = provider + provider_address # R7:AV7
    contract_shipments = '№ п/п 1 №579 от 12.07.2021 г.' # R10:AV10
    goods = 'Жмых Соевый, дробленный не для кормовых целей' # J15:R15
    goods_code = '168' # W15:X15
    goods_volume = '27,48' # AA15:AC15
    goods_cost_per_one = '56000' # AD15:AM15
    goods_cost_total_without_tax = '1538880' # AN15:AU15
    goods_cost_tax = '153888' # BB15:BE15
    goods_cost_total_with_tax = '1692768' # BF15:BI15
    total_cost_without_tax = '1538880' # AN16:AU16
    total_cost_tax = '153888' # BB16:BE16
    total_cost_with_tax = '1692768' # BF16:BI16
    contract = '№ 128/ДМА от 09.03.2021 (RUB)' # T22:BQ23
    text_for_C40_to_AL40 = '{0}, ИНН/КПП {1}'.format(provider, provider_inn_and_kpp) # C40:AL40

    ws['P1'] = invoice_number
    ws['Y1'] = invoice_date
    ws['R4'] = provider
    ws['R5'] = provider_address
    ws['R6'] = provider_inn_and_kpp
    ws['R7'] = text_for_R7_to_AV7
    ws['R10'] = contract_shipments
    ws['J15'] = goods
    ws['W15'] = goods_code
    ws['AA15'] = goods_volume
    ws['AD15'] = goods_cost_per_one
    ws['AN15'] = goods_cost_total_without_tax
    ws['BB15'] = goods_cost_tax
    ws['BF15'] = goods_cost_total_with_tax
    ws['AN16'] = total_cost_without_tax
    ws['BB16'] = total_cost_tax
    ws['BF16'] = total_cost_with_tax
    ws['T22'] = contract
    ws['C40'] = text_for_C40_to_AL40

    tpl_save_path = os.path.join(BASE_DIR, 'temp_file', 'temporary_files', 'УПД.xlsx')
    wb.save(tpl_save_path)
    save_doc_after_generation('УПД.xlsx', 'universal_transfer_document', tpl_save_path)


def gen_doc_letter_for_refund():
    tpl_path = os.path.join(BASE_DIR, 'temp_file' , 'template_doc', 'return_letter.docx')
    tpl = DocxTemplate(tpl_path)

    context = {
        'to': 'ИП Лесных В.Н.',
        'date': '20.07.2021 г.',
        'name_head': 'Виктор Николаевич',
        'conract': '№ 53/ДМ от 04.08.2020',
        'cost': '602888,00',
        'cost_with_words': 'Шестьсот две тысячи восемьсот восемьдесят восемь руб 00 коп',
        'correspondent_account': '30101810000000000272',
        'bank_name': 'ПАО БАНК ЗЕНИТ, г. Москва',
        'bik': '044525272',
        'payment_account': '40702810700002026956',
    }

    tpl.render(context)
    tpl_save_path = os.path.join(BASE_DIR, 'temp_file' , 'temporary_files', 'Письмо_на_возврат.docx')
    tpl.save(tpl_save_path)
    save_doc_after_generation('Письмо_на_возврат.docx', 'letter_for_refund', tpl_save_path)